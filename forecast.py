#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""

Optimierungspotenziale:

- Integriere Kundenanzahl
- Integriere Monthly Churn basierend auf Kundenanzahl


- Integriere NRR durch Preissteigerung und neue Abos
    () step 1 overall
    () step 2 granular je Maßnahme


- Integriere Conversion Rate, Terminierungsrate, Monthly Churn und NRR für jedes Paket und jede Lead-Quelle

- Integriere maximale Anzahl an Setter Terminen
- Integriere maximale verbundene Telfonate je Setter

- Integriere Saisonalität durch Monatsmultiplikator
- Integriere Zeit bis zum Closing



SaaS Vertriebs-Forecast – Setter/Closer-Skalierung (Mai 2027, 60 Monate)
========================================================================

BERECHNUNGSLOGIK (je Monat m, als Monatszahl 1, 2, 3, ... 60):

1) SETTER-TERMINE
   - Die Gesamtzahl der verbundenen Telefonate (CONNECTED_CALLS_TOTAL, aktuell 300)
     wird GLEICHMÄSSIG auf alle in Monat m bereits eingestellten Setter aufgeteilt.
   - Für jeden Setter wird die Differenz zwischen dem zu berechnenden Monat m und
     seinem Einstellungsmonat gebildet -> seine "Betriebszugehörigkeit" in Monaten
     (Einstellungsmonat selbst = Monat 1 der Zugehörigkeit).
   - Anhand dieser Differenz wird seine Terminquote bestimmt:
       * Zugehörigkeit 1..5  -> Ramp-Up-Wert SETTER_RAMP[Zugehörigkeit - 1]
       * Zugehörigkeit > 5   -> voller Zielwert SETTER_TARGET_RATE
   - Termine des Setters = (Telefonate je Setter) * (seine Terminquote).
   - Alle Setter-Termine werden summiert -> Termine gesamt des Monats.

2) CLOSER-ABSCHLÜSSE
   - Die Termine gesamt werden GLEICHMÄSSIG auf alle in Monat m bereits
     eingestellten Closer aufgeteilt.
   - Für jeden Closer wird auf demselben Weg (Monat m minus Einstellungsmonat)
     seine Closing Rate bestimmt:
       * Zugehörigkeit 1..5  -> Ramp-Up-Wert CLOSER_RAMP[Zugehörigkeit - 1]
       * Zugehörigkeit > 5   -> voller Zielwert CLOSER_TARGET_RATE
   - Abschlüsse des Closers = (Termine je Closer) * (seine Closing Rate).
   - Alle Abschlüsse werden summiert -> Abschlüsse gesamt des Monats.

3) MRR
   - MRR-Zuwachs des Monats = Abschlüsse gesamt * DEAL_MRR (Ø 333 EUR je Deal).
   - MRR des Folgemonats    = MRR des letzten Monats + MRR-Zuwachs des letzten
     Monats. Startpunkt: MRR in Monat 1 = START_MRR (50.000 EUR).

DATENSÄTZE:
   SETTER_HIRES und CLOSER_HIRES sind einfache Listen von Monatszahlen.
   Jede Zahl = Einstellungsmonat einer Person (1 = Mai 2027, 2 = Juni 2027, ...).
   Werte <= 0 bedeuten: vor dem Forecast-Start eingestellt (z. B. -4 = so weit
   zurück, dass die Person ab Monat 1 bereits voll auf Zielwert ist).
   Neue Mitarbeiter planen = einfach weitere Monatszahlen an die Liste anhängen.
"""

from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ----------------------------------------------------------------------
# PARAMETER (hier anpassen)
# ----------------------------------------------------------------------
N_MONTHS = 60                    # Forecast-Länge: Mai 2027 ... April 2032
START_YEAR, START_MONTH = 2027, 5

CONNECTED_CALLS_TOTAL = 300      # verbundene Telefonate je Monat (Pool, wird auf Setter verteilt)
DEAL_MRR = 333.0                 # Ø MRR je abgeschlossenem Deal in EUR
START_MRR = 50_000.0             # MRR zu Beginn von Monat 1

# Ramp-Up: Wert je Zugehörigkeitsmonat 1..5, danach gilt der Zielwert
SETTER_RAMP = [0.00, 0.00, 0.10, 0.15, 0.20]   # Terminquote auf verbundene Telefonate
SETTER_TARGET_RATE = 0.20                       # Zielwert ab Monat 6

CLOSER_RAMP = [0.00, 0.00, 0.15, 0.25, 0.35]   # Closing Rate auf Termine
CLOSER_TARGET_RATE = 0.40                       # Zielwert ab Monat 6

# Einstellungsmonate: Startteam = 2 Setter + 2 Closer, so weit zurück
# eingestellt (-4), dass sie ab Monat 1 voll auf Zielwert sind.
SETTER_HIRES = [-4, -4]          # z. B. [-4, -4, 7, 13] -> 2 neue Setter in Monat 7 und 13
CLOSER_HIRES = [-4, -4]          # z. B. [-4, -4, 7]     -> 1 neuer Closer in Monat 7


# ----------------------------------------------------------------------
# BERECHNUNG
# ----------------------------------------------------------------------
def rate_for_tenure(tenure: int, ramp: list[float], target: float) -> float:
    """Ramp-Up- oder Zielwert anhand der Zugehörigkeit in Monaten (1, 2, 3, ...)."""
    if tenure < 1:
        return 0.0                      # noch nicht eingestellt
    if tenure <= len(ramp):
        return ramp[tenure - 1]         # Ramp-Up-Phase, Monat 1..5
    return target                        # danach voller Zielwert


def month_label(m: int) -> str:
    """Monatszahl (1 = Mai 2027) -> 'Mai 2027' usw."""
    total = (START_YEAR * 12 + START_MONTH - 1) + (m - 1)
    y, mo = divmod(total, 12)
    names = ["Jan", "Feb", "Mär", "Apr", "Mai", "Jun",
             "Jul", "Aug", "Sep", "Okt", "Nov", "Dez"]
    return f"{names[mo]} {y}"


def calculate_forecast() -> list[dict]:
    rows = []
    mrr = START_MRR                                        # MRR zu Beginn Monat 1

    for m in range(1, N_MONTHS + 1):
        # --- Schritt 1: Setter-Termine ---
        active_setters = [h for h in SETTER_HIRES if h <= m]
        appointments = 0.0
        if active_setters:
            calls_per_setter = CONNECTED_CALLS_TOTAL / len(active_setters)
            for hire in active_setters:
                tenure = m - hire + 1                      # Differenz zum Einstellungsmonat
                quote = rate_for_tenure(tenure, SETTER_RAMP, SETTER_TARGET_RATE)
                appointments += calls_per_setter * quote

        # --- Schritt 2: Closer-Abschlüsse ---
        active_closers = [h for h in CLOSER_HIRES if h <= m]
        deals = 0.0
        if active_closers:
            appts_per_closer = appointments / len(active_closers)
            for hire in active_closers:
                tenure = m - hire + 1
                closing_rate = rate_for_tenure(tenure, CLOSER_RAMP, CLOSER_TARGET_RATE)
                deals += appts_per_closer * closing_rate

        # --- Schritt 3: MRR ---
        mrr_growth = deals * DEAL_MRR

        rows.append({
            "monat": m,
            "datum": month_label(m),
            "setter": len(active_setters),
            "closer": len(active_closers),
            "termine": appointments,
            "abschluesse": deals,
            "mrr": mrr,                                    # aktueller MRR (Monatsbeginn)
            "mrr_zuwachs": mrr_growth,                     # geplanter Zuwachs dieses Monats
        })

        # MRR des Folgemonats = letzter Monats-MRR + letzter Monatszuwachs
        mrr = mrr + mrr_growth

    return rows


# ----------------------------------------------------------------------
# AUSGABE: Terminal
# ----------------------------------------------------------------------
def print_forecast(rows: list[dict]) -> None:
    print(f"{'Mon':>3} | {'Datum':<9} | {'Set.':>4} | {'Clo.':>4} | "
          f"{'Termine':>8} | {'Deals':>7} | {'MRR aktuell':>13} | {'MRR-Zuwachs':>12}")
    print("-" * 82)
    for r in rows:
        print(f"{r['monat']:>3} | {r['datum']:<9} | {r['setter']:>4} | {r['closer']:>4} | "
              f"{r['termine']:>8.1f} | {r['abschluesse']:>7.2f} | "
              f"{r['mrr']:>11,.0f} € | {r['mrr_zuwachs']:>10,.0f} €")
    end_mrr = rows[-1]["mrr"] + rows[-1]["mrr_zuwachs"]
    print("-" * 82)
    print(f"MRR nach Monat {rows[-1]['monat']} ({rows[-1]['datum']}): {end_mrr:,.0f} €")


# ----------------------------------------------------------------------
# AUSGABE: Excel
# ----------------------------------------------------------------------
def write_excel(rows: list[dict], path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Forecast"

    bold = Font(name="Arial", bold=True, size=10)
    normal = Font(name="Arial", size=10)
    grey = Font(name="Arial", italic=True, size=9, color="595959")
    head_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    head_fill = PatternFill("solid", fgColor="1F4E79")
    param_fill = PatternFill("solid", fgColor="DDEBF7")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "SaaS Vertriebs-Forecast – berechnet durch forecast.py (Werte, keine Formeln)"
    ws["A1"].font = Font(name="Arial", bold=True, size=13)
    ws["A2"] = ("Alle Zahlen werden im Python-Skript berechnet. Parameter und "
                "Einstellungsmonate im Skript anpassen und neu ausführen.")
    ws["A2"].font = grey

    # Parameterblock (dokumentiert die Annahmen dieser Berechnung)
    params = [
        ("Verbundene Telefonate gesamt / Monat", CONNECTED_CALLS_TOTAL),
        ("Ø MRR je Abschluss (EUR)", DEAL_MRR),
        ("Start-MRR Monat 1 (EUR)", START_MRR),
        ("Setter-Ramp Monat 1–5", ", ".join(f"{v:.0%}" for v in SETTER_RAMP)),
        ("Setter-Zielquote ab Monat 6", f"{SETTER_TARGET_RATE:.0%}"),
        ("Closer-Ramp Monat 1–5", ", ".join(f"{v:.0%}" for v in CLOSER_RAMP)),
        ("Closer-Zielrate ab Monat 6", f"{CLOSER_TARGET_RATE:.0%}"),
        ("Setter-Einstellungsmonate", str(SETTER_HIRES)),
        ("Closer-Einstellungsmonate", str(CLOSER_HIRES)),
    ]
    for i, (label, value) in enumerate(params, start=4):
        ws.cell(row=i, column=1, value=label).font = normal
        ws.cell(row=i, column=1).fill = param_fill
        ws.cell(row=i, column=1).border = border
        c = ws.cell(row=i, column=2, value=value)
        c.font = normal
        c.border = border

    header_row = len(params) + 5
    headers = ["Monat #", "Datum", "Setter aktiv", "Closer aktiv",
               "Termine", "Abschlüsse", "MRR aktuell (€)", "MRR-Zuwachs (€)"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = head_font
        c.fill = head_fill
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, r in enumerate(rows, start=header_row + 1):
        values = [r["monat"], r["datum"], r["setter"], r["closer"],
                  round(r["termine"], 1), round(r["abschluesse"], 2),
                  round(r["mrr"], 0), round(r["mrr_zuwachs"], 0)]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.font = normal
            c.border = border
            if col == 5:
                c.number_format = "0.0"
            elif col == 6:
                c.number_format = "0.00"
            elif col >= 7:
                c.number_format = '#,##0 "€"'

    # Abschlusszeile: MRR nach dem letzten Monat
    last = header_row + len(rows) + 1
    ws.cell(row=last, column=2, value="MRR nach Monat 60").font = bold
    end = ws.cell(row=last, column=7, value=round(rows[-1]["mrr"] + rows[-1]["mrr_zuwachs"], 0))
    end.font = bold
    end.number_format = '#,##0 "€"'

    for col, w in zip("ABCDEFGH", [36, 10, 11, 11, 10, 11, 15, 15]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=3).coordinate

    wb.save(path)
    print(f"\nExcel gespeichert: {path}")


if __name__ == "__main__":
    forecast = calculate_forecast()
    print_forecast(forecast)
    write_excel(forecast, "SaaS_Forecast_Python.xlsx")